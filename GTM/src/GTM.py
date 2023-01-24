import math
import numpy

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)


MAX_ITERATIONS:int = 1000

INV_TAU:float = 1 / math.tau

noBasisFunction:int
noLatVarSample:int

s:int

dataDimensions:int
latentDimensions:int

cycles:int
llh:numpy.ndarray

inputData:numpy.matrix

FI:numpy.matrix

beta:numpy.float64
outputData:numpy.matrix

globalR:numpy.matrix
globalDIST:numpy.matrix


def Grid( rowCount:int, columnCount:int, count:numpy.ndarray ) -> numpy.matrix:
    grid:numpy.matrix = numpy.empty( ( int( rowCount ), int( columnCount ) ), dtype = numpy.float64 )

    index:numpy.ndarray = numpy.zeros( columnCount, dtype = numpy.float64 )

    gridIndex:int = 0

    for i in range( 0, int( rowCount * columnCount ) ):
        currDim:int = i % columnCount
        dist:numpy.float64 = 2 / ( count[ currDim ] - 1 )
        grid.flat[ gridIndex ] = -1 + dist * index[ currDim ]
        gridIndex += 1

        if currDim == 0:
            index[ currDim ] = index[ currDim ] + 1
        else:
            allZero:bool = True
            for j in range( 0, currDim ):
                if index[ j ] != 0:
                    allZero = False
                    break

            if allZero:
                index[ currDim ] = index[ currDim ] + 1

        if index[ currDim ] >= count[ currDim ]:
            index[ currDim ] = 0

    return grid

def Distance( M:numpy.matrix, N:numpy.matrix, Dist:numpy.matrix ):
    for i in range( 0, N.shape[ 0 ] ):
        for j in range( 0, M.shape[ 0 ] ):
            Dist[ i ][ j ] = 0

            for k in range( 0, N[ i ].size ):
                diff = M[ j ][ k ] - N[ i ][ k ]
                Dist[ i ][ j ] += diff * diff

def GetPrincipalComponents( inputData:numpy.matrix ):
    covariant:numpy.matrix = numpy.cov( inputData.T )
    eigenValues, eigenVectors = numpy.linalg.eig( covariant )
    
    idx = eigenValues.argsort()[ ::-1 ]
    eigenValues = eigenValues[ idx ]
    eigenVectors = eigenVectors[ :, idx ]

    return ( eigenValues, eigenVectors )

def SetInitialOutput( X:numpy.matrix ):
    eigenValues, eigenVectors = GetPrincipalComponents( inputData )

    latentDimEigenValue:numpy.float64 = eigenValues[ latentDimensions ]

    diff:int = eigenValues.size - latentDimensions
    eigenValues = numpy.delete( eigenValues, -diff )
    eigenVectors = numpy.delete( eigenVectors, -diff, 1 )

    eigenValues = numpy.emath.sqrt( eigenValues )
    A:numpy.matrix = numpy.matmul( eigenVectors, numpy.diag( eigenValues ) )
    
    M:numpy.matrix = numpy.matmul( FI.T, FI )

    W:numpy.matrix = numpy.matmul( numpy.matmul( numpy.linalg.inv( M ), FI.T ), numpy.matmul( X, A.T ) )
    W[ -1 ] = inputData.mean( 0 )[ 0 ]

    global outputData
    outputData = numpy.matmul( FI, W )

    interDist:numpy.matrix = numpy.empty( ( outputData.shape[ 0 ], outputData.shape[ 0 ] ), dtype = numpy.float64 )
    Distance( outputData, outputData, interDist )
    numpy.fill_diagonal( interDist, 0x7fffffff )

    global beta
    minColumns:numpy.matrix = interDist.min( 0 )
    beta = 2 / minColumns.mean()

    if latentDimensions < dataDimensions:
        beta = min( beta, 1 / latentDimEigenValue )

def Setup():
    count:numpy.ndarray = numpy.full( latentDimensions, noLatVarSample )
    total:int = count[ 0 ]
    for i in range( 1, count.size ):
        total *= count[ i ]
    X:numpy.matrix = Grid( total, latentDimensions, count )
    
    count:numpy.ndarray = numpy.full( latentDimensions, noBasisFunction )
    total:int = count[ 0 ]
    for i in range( 1, count.size ):
        total *= count[ i ]
    MU:numpy.matrix = Grid( total, latentDimensions, count )

    sigma:numpy.float64 = s * ( 2 / ( noBasisFunction - 1 ) )

    global FI
    FI = numpy.empty( ( X.shape[ 0 ], MU.shape[ 0 ] + 1 ), dtype = numpy.float64 )
    Distance( MU, X, FI )
    FI *= -1 / ( 2 * sigma * sigma )
    FI = numpy.exp( FI )
    FI[ :, -1 ] = numpy.ones( FI.shape[ 0 ] )

    SetInitialOutput( X )

    global globalR
    global globalDIST
    globalR = numpy.empty( ( FI.shape[ 0 ], inputData.shape[ 0 ] ), dtype = numpy.float64 )
    globalDIST = numpy.empty( ( FI.shape[ 0 ], inputData.shape[ 0 ]), dtype = numpy.float64 )
    Distance( inputData, outputData, globalR )

def Create2dData():
    global noLatVarSample
    noLatVarSample      = 20
    global latentDimensions
    latentDimensions    = 1
    global noBasisFunction
    noBasisFunction     = 5
    global s
    s                   = 2
    global dataDimensions
    dataDimensions      = 2

    global inputData
    inputData           = numpy.empty( ( 59, dataDimensions ), dtype = numpy.float64 )
    x:float             = 0.15
    for index in range( 0, 59 ):
        inputData[ index ][ 0 ] = x
        inputData[ index ][ 1 ] = x + 1.25 * math.sin( 2.0 * x )
        x += 0.05

def CalculateResponsibilities() -> numpy.float64:
    mul:numpy.float64 = -beta / 2
    logSum:numpy.float64 = 0

    invColumnSum:numpy.matrix = numpy.empty( inputData.shape[ 0 ] )

    for i in range( 0, globalR.shape[ 1 ] ):
        globalR[ 0 ][ i ] = math.exp( globalR[ 0 ][ i ] * mul )
        invColumnSum[ i ] = globalR[ 0 ][ i ]

    for j in range( 1, globalR.shape[ 0 ] - 1 ):
        for i in range( 0, globalR.shape[ 1 ] ):
            globalR[ j ][ i ] = math.exp( globalR[ j ][ i ] * mul )
            invColumnSum[ i ] += globalR[ j ][ i ]

    for i in range( 0, globalR.shape[ 1 ] ):
        globalR[ -1 ][ i ] = math.exp( globalR[ -1 ][ i ] * mul )
        invColumnSum[ i ] += globalR[ -1 ][ i ]
        logSum += math.log( invColumnSum[ i ] )
        invColumnSum[ i ] = 1 / invColumnSum[ i ]

    for j in range( 0, globalR.shape[ 0 ] ):
        globalR[ j ] *= invColumnSum

    return logSum

def Train():
    logSum:numpy.float64 = CalculateResponsibilities()
    
    global globalR
    global globalDIST
    global beta
    
    global llh
    llh[ cycles ] = ( inputData.shape[ 1 ] / 2 ) * math.log( beta * INV_TAU ) - math.log( noLatVarSample )
    llh[ cycles ] = logSum + globalR.shape[ 1 ] * llh[ cycles ]

    rowSums:numpy.matrix = globalR.sum( 1 )
    A:numpy.matrix = numpy.matmul( numpy.matmul( FI.T, numpy.diag( rowSums ) ), FI )
    W:numpy.matrix = numpy.matmul( numpy.linalg.inv( A ), numpy.matmul( FI.T, numpy.matmul( globalR, inputData ) ) )

    global outputData
    outputData = numpy.matmul( FI, W )

    Distance( inputData, outputData, globalDIST )

    totalSum:numpy.float64 = 0

    for i in range( 0, globalR.shape[ 0 ] ):
        for j in range( 0, globalR.shape[ 1 ] ):
            totalSum += globalR[ i ][ j ] * globalDIST[ i ][ j ]

    swap:numpy.matrix = globalR
    globalR = globalDIST
    globalDIST = swap

    beta = ( inputData.shape[ 0 ] * inputData.shape[ 1 ] ) / totalSum


Create2dData()
Setup()

cycles = 0
llh = numpy.empty( ( MAX_ITERATIONS ), dtype = numpy.float64 )
while cycles <= 1 or abs( llh[ cycles - 1 ] - llh[ cycles - 2 ] ) > 0.01:
    Train()
    cycles += 1

    if cycles >= MAX_ITERATIONS:
        break

workbook = Workbook()

sheet = workbook.active
sheet.title = "Data"

sheet[ "A1" ] = "Input"
sheet[ "A2" ] = "x"
sheet[ "B2" ] = "y"

for index in range( 0, inputData.shape[ 0 ] ):
    sheet.cell( index + 3, 1, inputData[ index ][ 0 ] )
    sheet.cell( index + 3, 2, inputData[ index ][ 1 ] )

sheet[ "C1" ] = "Output"
sheet[ "C2" ] = "x"
sheet[ "D2" ] = "y"

for index in range( 0, outputData.shape[ 0 ] ):
    sheet.cell( index + 3, 3, outputData[ index ][ 0 ] )
    sheet.cell( index + 3, 4, outputData[ index ][ 1 ] )


scatterChart = ScatterChart()

outputValuesX = Reference( sheet, min_col = 3, min_row = 3, max_col = 3, max_row = outputData.shape[ 0 ] + 3 )
outputValuesY = Reference( sheet, min_col = 4, min_row = 3, max_col = 4, max_row = outputData.shape[ 0 ] + 3 )
outputSeries = Series( outputValuesY, outputValuesX )
outputSeries.title = openpyxl.chart.series.SeriesLabel( openpyxl.chart.data_source.StrRef( "'Data'!C1" ) )
outputSeries.smooth = False;
outputSeries.marker.symbol = "circle"
outputSeries.marker.size = math.sqrt( 1 / beta ) * 100
outputSeries.graphicalProperties.solidFill = "BE4B48"
outputSeries.graphicalProperties.line.solidFill = "BE4B48"
scatterChart.series.append( outputSeries )

inputValuesX = Reference( sheet, min_col = 1, min_row = 3, max_col = 1, max_row = inputData.shape[ 0 ] + 3 )
inputValuesY = Reference( sheet, min_col = 2, min_row = 3, max_col = 2, max_row = inputData.shape[ 0 ] + 3 )
inputSeries = Series( inputValuesY, inputValuesX )
inputSeries.title = openpyxl.chart.series.SeriesLabel( openpyxl.chart.data_source.StrRef( "'Data'!A1" ) )
inputSeries.smooth = False;
inputSeries.marker.symbol = "circle"
inputSeries.marker.size = 20
inputSeries.graphicalProperties.solidFill = "2A6099"
inputSeries.graphicalProperties.line.solidFill = "2A6099"
scatterChart.series.append( inputSeries )

sheet.add_chart( scatterChart, "F3" )

workbook.save( "../GTM.xlsx" )